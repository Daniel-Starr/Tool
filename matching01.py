import pandas as pd
import re
from fuzzywuzzy import process, fuzz
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict
import logging
from typing import Dict, List, Set, Tuple
from dataclasses import dataclass
import tkinter as tk
from tkinter import ttk, messagebox

# 样式设置
STYLE_CONFIG = {
    "header_font": Font(name='微软雅黑', bold=True, color="FFFFFF"),
    "header_fill": PatternFill(start_color="4472C4", fill_type="solid"),
    "body_font": Font(name='宋体', size=10),
    "highlight_yellow": PatternFill(start_color="FFFF00", fill_type="solid"),
    "alignment": Alignment(vertical="center", wrap_text=True)
}

VOLTAGE_PATTERNS = [
    (r"(?:交流|直流)?\s*(\d+)\s*[kK][Vv]?", 1),
    (r"\b(\d+)\s*[kK][Vv]?\b", 1),
    (r"[\(（](.*?\d+\s*[kK][Vv]?)", 1),
    (r"([A-Za-z]+[-_]?\d+\s*[kK][Vv]?)", 1),
    (r"(\d+)\s*千伏", 1),
]

COLUMN_MAPPING = {
    "model_data": {
        "标识系统编码": ["电网工程标识系统编码", "SYSTEM_标识系统编码", "标识系统编码"],
        "工程名称": ["工程中名称", "SYSTEM_工程名称", "工程名称"],
        "电压等级": ["电压等级", "SYSTEM_电压等级"],
        "实物ID": ["SYSTEM_实物ID", "实物ID"],
        "Device_ID": ["Device_ID"]
    },
    "reference_data": {
        "电压等级": ["电压等级"],
        "设备类型": ["设备类型"],
        "设备名称": ["设备名称"],
        "实物ID": ["实物ID"]
    }
}


@dataclass
class MatchResult:
    system_code: str
    project_name: str
    model_voltage: str
    ref_voltage: str
    device_type: str
    physical_id: str
    match_status: str
    ref_device_name: str
    similarity: int


class DataProcessor:
    def __init__(self, model_path: str, ref_path: str):
        self.model_df = self._load_and_preprocess(model_path, "model_data")
        self.ref_df = self._load_and_preprocess(ref_path, "reference_data")
        self.voltage_groups = self._prepare_reference_data()

    def _load_and_preprocess(self, path: str, data_type: str) -> pd.DataFrame:
        try:
            df = pd.read_excel(path, engine='openpyxl')
            df = self._standardize_columns(df, data_type)
            df["电压等级"] = df["电压等级"].astype(str).fillna("Unknown")
            if data_type == "model_data":
                df["Device_ID"] = df.get("Device_ID", pd.NA)
            return df
        except FileNotFoundError:
            raise FileNotFoundError(f"文件未找到: {path}")
        except Exception as e:
            raise Exception(f"读取文件 {path} 时出错: {str(e)}")

    def _standardize_columns(self, df: pd.DataFrame, data_type: str) -> pd.DataFrame:
        column_mapping = {}
        for target, candidates in COLUMN_MAPPING[data_type].items():
            for candidate in candidates:
                if candidate in df.columns:
                    column_mapping[candidate] = target
                    break
            else:
                df[target] = pd.NA
        return df.rename(columns=column_mapping)

    def _prepare_reference_data(self) -> Dict[str, pd.DataFrame]:
        self.ref_df["参考电压"] = self.ref_df["电压等级"].apply(self.clean_voltage)
        self.ref_df["设备类型_清洗"] = (
            self.ref_df["设备类型"].astype(str)
            .str.replace(r"[\d#\s\-_]+", "", regex=True)
            .str.strip()
        )
        return {v: g for v, g in self.ref_df.groupby("参考电压") if v != "Unknown"}

    @staticmethod
    def clean_voltage(text: str) -> str:
        clean_text = str(text).strip().replace(" ", "").upper()
        for pattern, group_idx in VOLTAGE_PATTERNS:
            match = re.search(pattern, clean_text, flags=re.IGNORECASE)
            if match:
                voltage = match.group(group_idx).upper().replace("KV", "kV")
                return f"{voltage}kV" if not voltage.endswith("kV") else voltage
        return "Unknown"

    def match_devices(self) -> List[MatchResult]:
        results = []
        for _, row in self.model_df.iterrows():
            best = self._find_best_match(row)
            results.append(self._build_result(row, best))
        return results

    def _find_best_match(self, model_row: pd.Series) -> Dict:
        best = {"score": 0, "data": None, "voltage": "无匹配"}
        model_voltage = str(model_row.get("电压等级", ""))
        for ref_voltage, ref_group in self.voltage_groups.items():
            if not self._is_voltage_match(ref_voltage, model_voltage):
                continue
            for match, score in self._fuzzy_match(model_row["工程名称"], ref_group["设备名称"]):
                if score > best["score"]:
                    best.update({
                        "score": score,
                        "data": ref_group[ref_group["设备名称"] == match].iloc[0],
                        "voltage": ref_voltage
                    })
        return best

    @staticmethod
    def _fuzzy_match(query: str, choices: pd.Series) -> List[Tuple[str, int]]:
        if not isinstance(query, str):
            query = str(query) if not pd.isna(query) else ""

        # 确保 choices 是字符串类型且非空
        choices = choices.dropna().astype(str)
        choices_list = choices.tolist()

        if not choices_list or not query.strip():
            return []

        try:
            # 使用 process.extract 进行模糊匹配
            matches = process.extract(query, choices_list, scorer=fuzz.token_sort_ratio, limit=3)
            return [(match, score) for match, score in matches]
        except Exception as e:
            print(f"模糊匹配出错: {e}")
            return []

    @staticmethod
    def _is_voltage_match(ref_v: str, model_v: str) -> bool:
        try:
            ref_match = re.search(r"\d+", ref_v)
            model_match = re.search(r"\d+", model_v)
            if ref_match and model_match:
                return int(ref_match.group()) == int(model_match.group())
            return False
        except Exception:
            return False

    @staticmethod
    def _build_result(model_row: pd.Series, best: Dict) -> MatchResult:
        similarity = best["score"]
        status = "无匹配"

        if similarity >= 90:
            status = f"{similarity}% ★★★★"
        elif similarity >= 70:
            status = f"{similarity}% ★★★"
        elif similarity >= 50:
            status = f"{similarity}% ★★"
        elif similarity >= 30:
            status = f"{similarity}% ★"
        elif similarity > 0:
            status = f"{similarity}%"

        return MatchResult(
            system_code=str(model_row.get("标识系统编码", "")),
            project_name=str(model_row.get("工程名称", "")),
            model_voltage=str(model_row.get("电压等级", "")),
            ref_voltage=best.get("voltage", "无匹配"),
            device_type=best["data"]["设备类型_清洗"] if best["data"] is not None else "无匹配",
            physical_id=best["data"]["实物ID"] if best["data"] is not None else "需手动输入",
            match_status=status,
            ref_device_name=best["data"]["设备名称"] if best["data"] is not None else "",
            similarity=similarity
        )


def generate_initial_report(results: List[MatchResult], model_df: pd.DataFrame, filename: str):
    wb = Workbook()
    ws = wb.active
    ws.append(
        ["标识系统编码", "工程名称", "model电压", "参考电压", "参考设备类型", "实物ID", "匹配状态", "参考设备名称",
         "相似度", "Device_ID"])

    for i, r in enumerate(results):
        device_id = ""
        if i < len(model_df):
            device_id = str(model_df.iloc[i].get("Device_ID", ""))

        ws.append([
            r.system_code, r.project_name, r.model_voltage, r.ref_voltage,
            r.device_type, r.physical_id, r.match_status, r.ref_device_name,
            r.similarity, device_id
        ])

    try:
        wb.save(filename)
        print(f"✅ 报告已生成: {filename}")
    except Exception as e:
        print(f"❌ 保存报告失败: {e}")
        raise


def launch_manual_correction_gui(filepath: str, ref_df: pd.DataFrame):
    try:
        df = pd.read_excel(filepath)
        available_ids = sorted(ref_df["实物ID"].dropna().astype(str).unique())
    except Exception as e:
        print(f"❌ 启动GUI失败: {e}")
        return

    root = tk.Tk()
    root.title("实物ID手动校对工具")
    tree = ttk.Treeview(root, columns=list(df.columns), show="headings")
    for col in df.columns:
        tree.heading(col, text=col)
        tree.column(col, width=100)
    tree.pack(fill=tk.BOTH, expand=True)

    for idx, row in df.iterrows():
        tree.insert("", "end", iid=str(idx), values=list(row))

    def on_double_click(event):
        selected = tree.selection()
        if not selected:
            return
        idx = selected[0]
        popup = tk.Toplevel(root)
        popup.title("选择实物ID")
        cb = ttk.Combobox(popup, values=available_ids)
        cb.pack(pady=10)

        def save_selection():
            val = cb.get()
            if val:
                df.at[int(idx), "实物ID"] = val
                tree.set(idx, "实物ID", val)
            popup.destroy()

        ttk.Button(popup, text="确定", command=save_selection).pack()

    def export_file():
        try:
            df.to_excel("最终修正结果.xlsx", index=False)
            messagebox.showinfo("导出成功", "文件已保存为：最终修正结果.xlsx")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    tree.bind("<Double-1>", on_double_click)
    ttk.Button(root, text="导出结果", command=export_file).pack(pady=10)
    root.mainloop()


def main():
    try:
        processor = DataProcessor("device_data.xlsx", "test_work.xlsx")
        results = processor.match_devices()
        generate_initial_report(results, processor.model_df, "智能设备匹配报告-专业版.xlsx")
        print("✅ 自动匹配完成，已生成初步 Excel 文件。")
        launch_manual_correction_gui("智能设备匹配报告-专业版.xlsx", processor.ref_df)
    except Exception as e:
        print(f"❌ 主程序执行失败: {e}")


if __name__ == "__main__":
    main()